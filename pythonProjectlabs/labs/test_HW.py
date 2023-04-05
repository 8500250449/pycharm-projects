import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
from labs.pageobject import registerpage

@pytest.fixture()
def setup():
    print("run before every method")

def test_def2(setup):
    print("Hello world")

def test_registeropencart(setup):
    excelpath = "C:\\Users\\Administrator\\PycharmProjects\\pythonProject1\\opencart\\pythonr1.xlsx"
    workbook = openpyxl.load_workbook(excelpath)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows)
    for i in range(2, rows + 1):
        fname = sheet.cell(row=i, column=1).value
        lname = sheet.cell(row=i, column=2).value
        email = sheet.cell(row=i, column=3).value
        pword = sheet.cell(row=i, column=4).value
        driver=webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        rp=registerpage(driver)
        rp.clickonmyaccountlink()
        rp.clickonregisterlink()
        rp.enterthefirstname(fname)
        rp.enterthelastname(lname)
        time.sleep(5)
        rp.enteremail(email)
        rp.enterpassword(pword)
        rp.clickonnewsletter()
        time.sleep(5)
        rp.clickoncheckbox()
        time.sleep(5)
        rp.clickoncontinue()
        time.sleep(5)