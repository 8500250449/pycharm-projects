import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from opencart.pageobject import loginpage


class testcase_001:
    def helloworld(self):
        print("Hello world")

    def test_loginopencart(self):
        excelpath = "C:\\Users\\Administrator\\PycharmProjects\\pythonProject1\\opencart\\PYTHONL1.xlsx"
        workbook = openpyxl.load_workbook(excelpath)
        sheet = workbook.active
        rows = sheet.max_row
        cols = sheet.max_column
        print(rows)
        for i in range(2, rows + 1):
            uname = sheet.cell(row=i, column=1).value
            pword = sheet.cell(row=i, column=2).value
            driver=webdriver.Chrome(ChromeDriverManager().install())
            driver.get("https://demo.opencart.com/")
            lp=loginpage(driver)
            lp.clickonmyaccount()
            lp.clickonloginlink()
            lp.enteremail(uname)
            lp.password(pword)
            lp.clickonsubmit()
            time.sleep(5)







