import time
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import pytest
from opencart.pageobject import loginpage


@pytest.fixture()
def setup():
    print("this runs before every method")


def testdef1(setup):
    print("Hello World")


def test_loginopencart(setup):
    excelpath = "C:\\Users\\Administrator\\PycharmProjects\\pythonProject1\\opencart\\PYTHONL1.xlsx"
    workbook = openpyxl.load_workbook(excelpath)
    sheet = workbook.active
    rows = sheet.max_row
    cols = sheet.max_column
    print(rows)
    for i in range(2, rows + 1):
        uname = sheet.cell(row=i, column=1).value
        pword = sheet.cell(row=i, column=2).value
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get("https://demo.opencart.com/")
        lp = loginpage(driver)
        lp.clickonmyaccountlink()
        lp.clickonloginlink()
        time.sleep(5)
        lp.enteremail(uname)
        lp.enterpassword(pword)
        time.sleep(5)
        lp.clickonsubmit()
        time.sleep(5)
