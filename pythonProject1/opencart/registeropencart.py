import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
def registeropencart(firstname,lastname,email,password):
    driver=webdriver.Chrome(ChromeDriverManager().install())

    driver.get("https://demo.opencart.com/")
    driver.find_element(By.LINK_TEXT,"My Account").click()
    driver.find_element(By.LINK_TEXT,"Register").click()

    if driver.title=="Your Store":
        print("title is matched")
    else:
        print("title not matched")

    driver.find_element(By.ID,"input-firstname").send_keys(firstname)
    driver.find_element(By.ID,"input-lastname").send_keys(lastname)
    driver.find_element(By.ID,"input-email").send_keys(email)
    driver.find_element(By.ID,"input-password").send_keys(password)
    driver.find_element(By.ID,"input-newsletter-yes").click()
    driver.find_element(By.XPATH,"//input[@type='checkbox']").click()
    driver.find_element(By.XPATH,"//button[@type='submit']").click()

excelpath = "C:\\Users\\Administrator\\PycharmProjects\\pythonProject1\\opencart\\pythonr1.xlsx"
workbook = openpyxl.load_workbook(excelpath)
sheet = workbook.active
rows = sheet.max_row
cols = sheet.max_column
print(rows)
for i in range(2, rows+1):
    fname=sheet.cell(row=i,column=1)  .value
    lname=sheet.cell(row=i, column=2).value
    email=sheet.cell(row=i, column=3).value
    pword=sheet.cell(row=i, column=4).value
    registeropencart(fname,lname,email,pword)



