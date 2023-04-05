import pytest
import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import  ChromeDriverManager
class testcase_001:
    def helloworld(self):
        print("Hello world")
    def test_loginopencart(self):

      driver=webdriver.Chrome(ChromeDriverManager().install())
      driver.get("https://demo.opencart.com/")
      driver.find_element(By.LINK_TEXT,"My Account").click()
      driver.find_element(By.LINK_TEXT,"Login").click()

    if driver.title=="Your Store":
        print("title is matched")
    else:
        print("title not matched")

    driver.find_element(By.ID,"input-email").send_keys(username)
    driver.find_element(By.ID,"input-password").send_keys(password)
    driver.find_element(By.XPATH,"//button[@type='submit']").click()
    driver.quit()

excelpath="C:\\Users\\Administrator\\PycharmProjects\\pythonProject1\\opencart\\PYTHONL1.xlsx"
workbook=openpyxl.load_workbook(excelpath)
sheet=workbook.active
rows=sheet.max_row
cols=sheet.max_column
print(rows)
for i in range(2,rows+1):
    uname=sheet.cell(row=i,column=1).value
    pword=sheet.cell(row=i,column=2).value
    loginopencart(uname,pword)